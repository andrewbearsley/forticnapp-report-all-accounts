#!/usr/bin/env python3
"""
Script to get consolidated compliance reports for all accounts in a Lacework instance.
Usage: python3 forticnapp_get_consolidated_report.py <api-key-path> <report-name> [options]
"""

import argparse
import csv
import datetime
import json
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================================
# Constants
# ============================================================================

class CloudType(Enum):
    """Supported cloud types."""
    AWS = 'aws'
    AZURE = 'azure'
    GCP = 'gcp'


class OutputFormat(Enum):
    """Supported output formats."""
    JSON = 'json'
    EXCEL = 'excel'


@dataclass
class Config:
    """Configuration constants."""
    # Rate limiting
    MAX_RETRIES: int = 5
    BACKOFF_INCREMENT: int = 30  # seconds (backoff delay when rate limited)
    
    # Request delays
    REQUEST_DELAY: float = 0.5  # seconds between requests (small delay to avoid hammering API)
    TAG_FETCH_DELAY: float = 2.0  # seconds between tag-fetch batches (higher to avoid 429 cascade)
    
    # Directories
    CACHE_DIR: str = 'cache'
    HISTORY_DIR: str = 'history'
    OUTPUT_DIR: str = 'output'
    
    # Default output files
    DEFAULT_EXCEL_OUTPUT: str = 'forticnapp-compliance-report.xlsx'
    DEFAULT_JSON_OUTPUT: str = 'forticnapp-compliance-report.json'

    # Policy documentation URL template ({policy_id} placeholder, uppercased with hyphens→underscores)
    POLICY_DOCS_URL: str = 'https://docs.fortinet.com/document/lacework-forticnapp/latest/lacework-forticnapp-policies?cshid={policy_id}'
    
    # Excel formatting
    EXCEL_HEADER_COLOR: str = "366092"
    EXCEL_HEADER_TEXT_COLOR: str = "FFFFFF"
    EXCEL_LINK_COLOR: str = "0000FF"
    EXCEL_MAX_COLUMN_WIDTH: int = 50
    
    # Severity mapping
    SEVERITY_MAP: Dict[int, str] = None
    SEVERITY_ORDER: Dict[str, int] = None
    
    # Summary field mappings
    RECOMMENDATIONS_FIELDS: List[Tuple[str, str]] = None
    POLICY_FIELDS: List[Tuple[str, str]] = None
    RESOURCE_FIELDS: List[Tuple[str, str]] = None
    SUMMARY_FIELDS: List[Tuple[str, str]] = None
    
    # Excel column headers
    EXCEL_HEADERS: List[str] = None
    
    def __post_init__(self):
        if self.SEVERITY_MAP is None:
            self.SEVERITY_MAP = {
                1: 'Critical',
                2: 'High',
                3: 'Medium',
                4: 'Low',
                5: 'Info'
            }
        if self.SEVERITY_ORDER is None:
            self.SEVERITY_ORDER = {
                'Critical': 1,
                'High': 2,
                'Medium': 3,
                'Low': 4,
                'Info': 5
            }
        # Reverse mapping: Policies API severity strings -> report format numbers
        self.SEVERITY_STRING_TO_NUM = {
            'critical': 1, 'high': 2, 'medium': 3, 'low': 4, 'info': 5
        }
        if self.RECOMMENDATIONS_FIELDS is None:
            # Policies overview section (total count)
            self.RECOMMENDATIONS_FIELDS = [
                ('NUM_RECOMMENDATIONS', 'Total Policies'),
            ]
        if self.POLICY_FIELDS is None:
            # Policy metrics (compliance recommendations/policies)
            self.POLICY_FIELDS = [
                ('NUM_COMPLIANT', 'Compliant'),
                ('NUM_NOT_COMPLIANT', 'Non-Compliant'),
                ('NUM_SEVERITY_1_NON_COMPLIANCE', 'Severity 1 (Critical)'),
                ('NUM_SEVERITY_2_NON_COMPLIANCE', 'Severity 2 (High)'),
                ('NUM_SEVERITY_3_NON_COMPLIANCE', 'Severity 3 (Medium)'),
                ('NUM_SEVERITY_4_NON_COMPLIANCE', 'Severity 4 (Low)'),
                ('NUM_SEVERITY_5_NON_COMPLIANCE', 'Severity 5 (Info)'),
                ('NUM_SUPPRESSED', 'Suppressed'),
            ]
        if self.RESOURCE_FIELDS is None:
            # Resource metrics (actual cloud resources)
            self.RESOURCE_FIELDS = [
                ('ASSESSED_RESOURCE_COUNT', 'Assessed'),
                ('VIOLATED_RESOURCE_COUNT', 'Violated'),
                ('SUPPRESSED_RESOURCE_COUNT', 'Suppressed'),
            ]
        if self.SUMMARY_FIELDS is None:
            # Combined for backward compatibility
            self.SUMMARY_FIELDS = self.RECOMMENDATIONS_FIELDS + self.POLICY_FIELDS + self.RESOURCE_FIELDS
        if self.EXCEL_HEADERS is None:
            self.EXCEL_HEADERS = [
                'Section', 'Policy', 'Severity', 'Account', 'Account Name',
                'Status', 'Resource', 'First Seen', 'Remediation', 'Docs', 'Tags'
            ]


# Global config instance
CONFIG = Config()

# Rate limit patterns
RATE_LIMIT_PATTERNS = [
    r'HTTP.*429',
    r'status.*429',
    r'429.*Too Many',
    r'rate limit exceeded',
    r'rate\.limit\.exceeded',
    r'too many requests'
]


# ============================================================================
# Output Utilities
# ============================================================================

class Colors:
    """ANSI color codes for terminal output."""
    RED = '\033[0;31m'
    GREEN = '\033[0;32m'
    YELLOW = '\033[1;33m'
    NC = '\033[0m'  # No Color


class Logger:
    """Centralized logging utility."""
    
    @staticmethod
    def info(msg: str) -> None:
        """Print info message."""
        print(f"{Colors.GREEN}[INFO]{Colors.NC} {msg}")

    @staticmethod
    def warning(msg: str) -> None:
        """Print warning message."""
        print(f"{Colors.YELLOW}[WARNING]{Colors.NC} {msg}", file=sys.stderr)

    @staticmethod
    def error(msg: str) -> None:
        """Print error message."""
        print(f"{Colors.RED}[ERROR]{Colors.NC} {msg}", file=sys.stderr)

    @staticmethod
    def verbose(msg: str, enabled: bool = False) -> None:
        """Print verbose message if enabled."""
        if enabled:
            print(f"{Colors.GREEN}[VERBOSE]{Colors.NC} {msg}", file=sys.stderr)


# ============================================================================
# Validation and Setup
# ============================================================================

def check_required_tools(output_format: OutputFormat) -> None:
    """Check if required tools are installed."""
    if not shutil.which("lacework"):
        Logger.error("lacework CLI is not installed. Please install it first.")
        sys.exit(1)

    if output_format == OutputFormat.EXCEL and not HAS_OPENPYXL:
        Logger.error("openpyxl is required for Excel output. Install it with: pip3 install openpyxl")
        sys.exit(1)


def load_api_key(api_key_path: str) -> Dict:
    """Load and validate API key from JSON file."""
    if not os.path.exists(api_key_path):
        Logger.error(f"API key file not found: {api_key_path}")
        sys.exit(1)

    with open(api_key_path, 'r') as f:
        api_key = json.load(f)
    
    required_fields = ['keyId', 'secret', 'account']
    missing = [field for field in required_fields if not api_key.get(field)]
    if missing:
        Logger.error(f"Invalid API key file. Missing required fields: {', '.join(missing)}")
        sys.exit(1)

    return api_key


def configure_lacework(api_key: Dict, verbose: bool) -> Dict[str, str]:
    """Configure Lacework CLI with API key and return environment variables."""
    Logger.verbose(f"Configuring Lacework CLI for account: {api_key['account']}", verbose)
    
    env = os.environ.copy()
    env['LW_ACCOUNT'] = api_key['account']
    env['LW_API_KEY'] = api_key['keyId']
    env['LW_API_SECRET'] = api_key['secret']
    
    if api_key.get('subAccount'):
        env['LW_SUBACCOUNT'] = api_key['subAccount']
        Logger.verbose(f"Using sub-account: {api_key['subAccount']}", verbose)
    
    # Test connection
    Logger.verbose("Testing Lacework CLI connection...", verbose)
    result = subprocess.run(
        ['lacework', 'configure', 'list'],
        capture_output=True,
        text=True,
        env=env
    )
    
    if result.returncode != 0:
        Logger.error("Failed to configure Lacework CLI. Please check your API key.")
        sys.exit(1)

    Logger.info("Successfully configured Lacework CLI")
    return env


# ============================================================================
# API Call Handling
# ============================================================================

def make_api_call(
    cmd: List[str],
    env: Dict[str, str],
    verbose: bool,
    retry_count: int = 0
) -> Tuple[str, bool]:
    """
    Make API call with rate limit handling.

    Returns:
        Tuple of (output, was_rate_limited). was_rate_limited is True when
        retries were exhausted due to rate limiting (output will be "").
    """
    while retry_count < CONFIG.MAX_RETRIES:
        Logger.verbose(f"Executing: {' '.join(cmd)}", verbose)
        
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                env=env
            )
            
            output = result.stdout + result.stderr
            is_rate_limited = _check_rate_limit(output, result.returncode)
            
            if is_rate_limited:
                _handle_rate_limit(retry_count, verbose, output)
                retry_count += 1
                continue
            
            if result.returncode != 0:
                # Check if output is valid JSON despite non-zero exit
                if _is_valid_json(output):
                    return output, False
                Logger.warning(f"Command failed with exit code {result.returncode}")
                Logger.verbose(f"Output: {output}", verbose)
                return "", False

            return output, False

        except Exception as e:
            Logger.warning(f"Error executing command: {e}")
            return "", False

    Logger.warning(f"Failed after {CONFIG.MAX_RETRIES} attempts")
    return "", True


def _check_rate_limit(output: str, exit_code: int) -> bool:
    """Check if output indicates rate limiting.

    Only checks error responses — valid JSON is never treated as rate-limited,
    since large payloads (e.g. /api/v2/Policies) can contain false matches
    like policy URLs with '1429' after 'HTTP'.
    """
    if exit_code == 0 and _is_valid_json(output):
        return False

    for pattern in RATE_LIMIT_PATTERNS:
        if re.search(pattern, output, re.IGNORECASE):
            return True

    return False


def _is_valid_json(text: str) -> bool:
    """Check if text is valid JSON."""
    try:
        json.loads(text)
        return True
    except (json.JSONDecodeError, ValueError):
        return False


def _handle_rate_limit(retry_count: int, verbose: bool, output: str) -> None:
    """Handle rate limit with incremental backoff."""
    if verbose:
        Logger.verbose("Received 429 response. Output was:", verbose)
        for line in output.split('\n'):
            if line.strip():
                Logger.verbose(f"  {line}", verbose)
    
    delay = CONFIG.BACKOFF_INCREMENT * (retry_count + 1)
    Logger.warning(
        f"Rate limited (429). Waiting {delay} seconds before "
        f"retry {retry_count + 1}/{CONFIG.MAX_RETRIES}..."
    )
    time.sleep(delay)


# ============================================================================
# Report Information
# ============================================================================

def get_report_info(report_name: str, env: Dict[str, str], verbose: bool) -> Optional[CloudType]:
    """Validate report name and determine cloud type."""
    Logger.verbose(f"Getting report definitions to validate: {report_name}", verbose)
    
    cmd = ['lacework', 'report-definitions', 'list', '--json']
    output, _ = make_api_call(cmd, env, verbose)
    
    try:
        definitions = json.loads(output)
        reports = _extract_reports_list(definitions)
        
        for report in reports:
            if report.get('reportName') == report_name:
                cloud_type = _determine_cloud_type(report.get('subReportType', ''))
                if cloud_type:
                    return cloud_type
        
        Logger.warning(f"Report '{report_name}' not found in report-definitions.")
        return None

    except json.JSONDecodeError:
        Logger.warning(f"Failed to parse report definitions response.")
        return None


def _extract_reports_list(definitions) -> List[Dict]:
    """Extract reports list from API response."""
    if isinstance(definitions, list):
        return definitions
    elif isinstance(definitions, dict) and 'data' in definitions:
        return definitions['data']
    return []


def _determine_cloud_type(sub_report_type: str) -> Optional[CloudType]:
    """Determine cloud type from sub-report type."""
    sub_type_lower = sub_report_type.lower()
    if sub_type_lower == 'aws':
        return CloudType.AWS
    elif sub_type_lower == 'azure':
        return CloudType.AZURE
    elif sub_type_lower in ['gcp', 'google']:
        return CloudType.GCP
    return None


# ============================================================================
# Account Fetching (with caching)
# ============================================================================

class AccountFetcher:
    """Base class for fetching cloud accounts with caching support."""
    
    def __init__(self, cloud_type: CloudType, env: Dict[str, str], verbose: bool, cache_dir: str):
        self.cloud_type = cloud_type
        self.env = env
        self.verbose = verbose
        self.cache_dir = cache_dir
        self.cache_file = os.path.join(cache_dir, f"{cloud_type.value}_accounts.txt")
    
    def get_accounts(self, use_cache: bool) -> List[str]:
        """Get accounts, using cache if requested."""
        if use_cache and os.path.exists(self.cache_file):
            Logger.verbose(f"Using cached {self.cloud_type.value} accounts", self.verbose)
            return self._read_cache()
        
        accounts = self._fetch_accounts()
        self._write_cache(accounts)
        return accounts
    
    def _read_cache(self) -> List[str]:
        """Read accounts from cache file."""
        with open(self.cache_file, 'r') as f:
            return [line.strip() for line in f if line.strip()]
    
    def _write_cache(self, accounts: List[str]) -> None:
        """Write accounts to cache file."""
        os.makedirs(self.cache_dir, exist_ok=True)
        with open(self.cache_file, 'w') as f:
            for account in accounts:
                f.write(f"{account}\n")
    
    def _fetch_accounts(self) -> List[str]:
        """Fetch accounts from API. Must be implemented by subclasses."""
        raise NotImplementedError
    
    def _get_cloud_account_statuses(self) -> Dict[str, Dict]:
        """Get cloud account integration statuses for all cloud types."""
        try:
            cmd = ['lacework', 'cloud-account', 'list', '--json']
            output, _ = make_api_call(cmd, self.env, self.verbose)
            try:
                data = json.loads(output)
            except json.JSONDecodeError:
                return {}
            
            # Determine integration type and account identifier key based on cloud type
            integration_type_map = {
                CloudType.AWS: 'AwsCfg',
                CloudType.AZURE: 'AzureCfg',
                CloudType.GCP: 'GcpCfg',
            }
            
            account_id_key_map = {
                CloudType.AWS: 'awsAccountId',
                CloudType.AZURE: 'tenantId',  # Azure uses tenantId (or tenantGuid), we'll map tenant/subscription
                CloudType.GCP: 'projectId',  # GCP uses projectId, we'll map org/project
            }
            
            integration_type = integration_type_map.get(self.cloud_type)
            account_id_key = account_id_key_map.get(self.cloud_type)
            
            if not integration_type or not account_id_key:
                return {}
            
            # Filter for the appropriate integration type and map by account identifier
            statuses = {}
            for acc in data:
                if acc.get('type') == integration_type:
                    account_id = acc.get('data', {}).get(account_id_key)
                    # For Azure, also check tenantGuid if tenantId is not available
                    if not account_id and self.cloud_type == CloudType.AZURE:
                        account_id = acc.get('data', {}).get('tenantGuid')
                    if account_id:
                        # Store by the base identifier and let subclasses handle mapping
                        statuses[account_id] = {
                            'ok': acc.get('state', {}).get('ok', None),
                            'enabled': acc.get('enabled', 0),
                            'name': acc.get('name', ''),
                            'intgGuid': acc.get('intgGuid', ''),
                            'details': acc.get('state', {}).get('details', {}),
                            'data': acc.get('data', {})  # Store full data for mapping
                        }
            return statuses
        except Exception as e:
            Logger.verbose(f"Could not fetch cloud account statuses: {e}", self.verbose)
            return {}
    
    def _log_account_statuses(self, accounts: List[Dict], id_key: str, status_key: str, cloud_account_statuses: Dict[str, Dict] = None, skipped_accounts: List[Tuple[str, str]] = None) -> None:
        """Log account statuses for verbose output."""
        skipped_ids = {acc_id for acc_id, _ in (skipped_accounts or [])}
        
        Logger.verbose("Account statuses:", self.verbose)
        
        disabled_accounts = []
        enabled_accounts = []
        
        for acc in accounts:
            account_id = acc.get(id_key, 'Unknown')
            status = acc.get(status_key, 'Unknown')
            
            # Skip accounts that are already being skipped due to integration errors
            if account_id in skipped_ids:
                continue
            
            # Log account with all available details
            if self.verbose:
                details = [f"status={status}"]
                if cloud_account_statuses and account_id in cloud_account_statuses:
                    cloud_status = cloud_account_statuses[account_id]
                    intg_ok = "OK" if cloud_status.get('ok') else "ERROR"
                    details.append(f"integration={intg_ok}")
                    if cloud_status.get('name'):
                        details.append(f"name={cloud_status.get('name')}")
                if acc.get('state'):
                    details.append(f"state={acc.get('state')}")
                if acc.get('integration_state'):
                    details.append(f"integration_state={acc.get('integration_state')}")
                if acc.get('account_alias'):
                    details.append(f"alias={acc.get('account_alias')}")
                Logger.verbose(f"  {account_id}: {', '.join(details)}", self.verbose)
            
            # Categorize accounts
            if status == 'Disabled':
                disabled_accounts.append(account_id)
            elif status == 'Enabled':
                enabled_accounts.append(account_id)
        
        if disabled_accounts:
            Logger.verbose(f"Skipping {len(disabled_accounts)} disabled account(s)", self.verbose)
        
        Logger.verbose(f"Processing {len(enabled_accounts)} enabled account(s)", self.verbose)


class AWSAccountFetcher(AccountFetcher):
    """Fetcher for AWS accounts."""
    
    def _fetch_accounts(self) -> List[str]:
        """Fetch AWS accounts from AwsCfg integrations."""
        Logger.verbose("Fetching AWS accounts from cloud-account integrations...", self.verbose)
        
        # Get AwsCfg integrations with state.ok == True
        cloud_account_statuses = self._get_cloud_account_statuses()
        
        accounts = []
        skipped_accounts = []
        found_integrations = []
        
        # Get compliance account statuses to verify accounts are enabled
        try:
            cmd = ['lacework', 'compliance', 'aws', 'list-accounts', '--json']
            output, _ = make_api_call(cmd, self.env, self.verbose)
            data = json.loads(output)
            compliance_accounts = {acc.get('account_id'): acc.get('status', '') 
                                   for acc in data.get('aws_accounts', [])}
        except (json.JSONDecodeError, KeyError):
            compliance_accounts = {}
        
        # Use integrations as primary source
        for account_id, status_info in cloud_account_statuses.items():
            integration_name = status_info.get('name', 'Unknown')
            # Only include integrations with state.ok == True
            if status_info.get('ok') is not True:
                skipped_accounts.append((account_id, integration_name))
                continue
            
            found_integrations.append((account_id, integration_name))
            
            # Verify account is enabled in compliance system
            compliance_status = compliance_accounts.get(account_id, '')
            if compliance_status != 'Enabled':
                Logger.verbose(f"Skipping {account_id}: not enabled in compliance system (status: {compliance_status})", self.verbose)
                continue
            
            accounts.append(account_id)
        
        # Log found integrations
        if found_integrations:
            Logger.info(f"Found {len(found_integrations)} AwsCfg integration(s) with state.ok == True:")
            for account_id, name in found_integrations:
                Logger.info(f"  {name} (account: {account_id})")
        
        # Log skipped accounts
        if skipped_accounts:
            Logger.warning(f"Found {len(skipped_accounts)} account(s) with integration errors:")
            for account_id, name in skipped_accounts:
                Logger.warning(f"  {name} (account: {account_id}): AwsCfg integration error (state.ok != True)")
            Logger.warning("These accounts have integration issues and have been skipped.")
        
        if not accounts:
            Logger.warning("No AWS accounts found with AwsCfg integrations (state.ok == True)")
        
        return sorted(set(accounts))


class AzureAccountFetcher(AccountFetcher):
    """Fetcher for Azure subscriptions."""
    
    def _fetch_accounts(self) -> List[str]:
        """Fetch Azure subscriptions from AzureCfg integrations."""
        Logger.verbose("Fetching Azure subscriptions from cloud-account integrations...", self.verbose)
        
        # Get AzureCfg integrations with state.ok == True
        cloud_account_statuses = self._get_cloud_account_statuses()
        
        # Get tenant IDs from integrations (only those with state.ok == True)
        tenant_ids_from_integrations = []
        skipped_integrations = []
        found_integrations = []
        
        for tenant_id, status_info in cloud_account_statuses.items():
            integration_name = status_info.get('name', 'Unknown')
            if status_info.get('ok') is True:
                tenant_ids_from_integrations.append(tenant_id)
                found_integrations.append((tenant_id, integration_name))
            else:
                skipped_integrations.append((tenant_id, integration_name))
        
        # Log found integrations
        if found_integrations:
            Logger.info(f"Found {len(found_integrations)} AzureCfg integration(s) with state.ok == True:")
            for tenant_id, name in found_integrations:
                Logger.info(f"  {name} (tenant: {tenant_id})")
        
        # Log skipped integrations
        if skipped_integrations:
            Logger.warning(f"Found {len(skipped_integrations)} AzureCfg integration(s) with errors:")
            for tenant_id, name in skipped_integrations:
                Logger.warning(f"  {name} (tenant: {tenant_id}): AzureCfg integration error (state.ok != True)")
            Logger.warning("These integrations have issues and have been skipped.")
        
        # Try list-tenants command
        cmd = ['lacework', 'compliance', 'azure', 'list-tenants', '--json']
        output, _ = make_api_call(cmd, self.env, self.verbose)
        
        tenants = []
        try:
            data = json.loads(output)
            # Handle different response structures
            if 'azure_tenants' in data and data['azure_tenants']:
                tenants = data.get('azure_tenants', [])
            elif 'azure_subscriptions' in data and data['azure_subscriptions']:
                # If list-tenants returns subscriptions directly, use them
                # But we still need tenant_id, so use from integrations
                subscriptions = data.get('azure_subscriptions', [])
                if subscriptions and tenant_ids_from_integrations:
                    # Use first tenant ID from integrations
                    tenant_id = tenant_ids_from_integrations[0]
                    tenants = [{'tenant_id': tenant_id}]
        except (json.JSONDecodeError, KeyError):
            pass
        
        # Use tenant IDs from cloud-account integrations (primary source)
        if not tenants and tenant_ids_from_integrations:
            Logger.verbose(f"Using tenant IDs from cloud-account integrations: {tenant_ids_from_integrations}", self.verbose)
            tenants = [{'tenant_id': tid} for tid in tenant_ids_from_integrations]
        elif tenants:
            # Filter tenants to only those from our integrations
            tenants = [t for t in tenants if t.get('tenant_id') in tenant_ids_from_integrations]
        
        all_subscriptions = []
        subscriptions_json = []
        
        for tenant in tenants:
            tenant_id = tenant.get('tenant_id')
            if not tenant_id:
                continue
            
            cmd = ['lacework', 'compliance', 'azure', 'list-subscriptions', tenant_id, '--json']
            sub_output, _ = make_api_call(cmd, self.env, self.verbose)
            
            subscriptions = []
            if not sub_output or not sub_output.strip():
                # Empty response means no subscriptions configured
                Logger.verbose(f"No subscriptions found for tenant {tenant_id}", self.verbose)
            else:
                try:
                    sub_data = json.loads(sub_output)
                    if isinstance(sub_data, list):
                        # Response is a list — may contain objects with nested 'subscriptions' arrays
                        # e.g., [{"subscriptions": [{"id": "...", "alias": "..."}]}]
                        subscriptions = []
                        for item in sub_data:
                            if isinstance(item, dict) and 'subscriptions' in item:
                                subscriptions.extend(item['subscriptions'])
                            elif isinstance(item, dict):
                                subscriptions.append(item)
                        # If no nested subscriptions found, use the list as-is
                        if not subscriptions:
                            subscriptions = sub_data
                    elif isinstance(sub_data, dict):
                        subscriptions = sub_data.get('azure_subscriptions', [])
                    else:
                        subscriptions = []
                except json.JSONDecodeError as e:
                    Logger.verbose(f"Could not parse subscriptions JSON for tenant {tenant_id}: {e}", self.verbose)
                    subscriptions = []

                for sub in subscriptions:
                    # Handle both dict and string formats
                    if isinstance(sub, dict):
                        sub_id = sub.get('subscription_id') or sub.get('id')
                        status = sub.get('status', 'Enabled')
                    else:
                        # If subscription is just an ID string
                        sub_id = str(sub)
                        status = 'Enabled'
                    
                    if sub_id:
                        # Store subscription info with tenant_id for status checking
                        sub_info = {
                            'tenant_id': tenant_id,
                            'subscription_id': sub_id,
                            'status': status,
                            'account_key': f"{tenant_id}/{sub_id}"  # Format used in account list
                        }
                        subscriptions_json.append(sub_info)
                        all_subscriptions.append(sub_info)
            
        # Filter subscriptions: enabled only (integration status already checked)
        accounts = []
        
        for sub_info in all_subscriptions:
            status = sub_info['status']
            account_key = sub_info['account_key']
            
            # Skip disabled subscriptions
            if status != 'Enabled':
                continue
            
            # Integration status already checked (we only process tenants with state.ok == True)
            accounts.append(account_key)
        
        # Log account statuses
        self._log_account_statuses(subscriptions_json, 'account_key', 'status', cloud_account_statuses, None)
        
        if not accounts:
            if tenant_ids_from_integrations:
                Logger.warning(f"No Azure subscriptions found for tenant(s): {', '.join(tenant_ids_from_integrations)}")
                Logger.warning("Subscriptions may need to be discovered/enabled in Lacework for this tenant.")
            else:
                Logger.warning("No Azure subscriptions found. No AzureCfg integrations with state.ok == True found.")
        
        return sorted(set(accounts))


class GCPAccountFetcher(AccountFetcher):
    """Fetcher for GCP projects."""
    
    def _fetch_accounts(self) -> List[str]:
        """Fetch GCP projects from GcpCfg integrations."""
        Logger.verbose("Fetching GCP projects from cloud-account integrations...", self.verbose)
        
        # Get GcpCfg integrations with state.ok == True
        cloud_account_statuses = self._get_cloud_account_statuses()
        
        accounts = []
        skipped_accounts = []
        found_integrations = []
        
        # Get compliance project statuses to verify projects are enabled
        try:
            cmd = ['lacework', 'compliance', 'google', 'list', '--json']
            output, _ = make_api_call(cmd, self.env, self.verbose)
            data = json.loads(output)
            # Build a map of project_id -> (org_id, status)
            compliance_projects = {}
            for proj in data.get('gcp_projects', []):
                proj_id = proj.get('project_id', '')
                org_id = proj.get('organization_id', 'n/a')
                status = proj.get('status', '')
                if proj_id:
                    compliance_projects[proj_id] = (org_id, status)
        except (json.JSONDecodeError, KeyError):
            compliance_projects = {}
        
        # Use integrations as primary source
        for project_id, status_info in cloud_account_statuses.items():
            integration_name = status_info.get('name', 'Unknown')
            # Only include integrations with state.ok == True
            if status_info.get('ok') is not True:
                skipped_accounts.append((project_id, integration_name))
                continue
            
            found_integrations.append((project_id, integration_name))
            
            # Verify project is enabled in compliance system
            if project_id in compliance_projects:
                org_id, compliance_status = compliance_projects[project_id]
                if compliance_status != 'Enabled':
                    Logger.verbose(f"Skipping {project_id}: not enabled in compliance system (status: {compliance_status})", self.verbose)
                    continue
                
                # Build account key
                if org_id == 'n/a':
                    account_key = f"n/a/{project_id}"
                else:
                    account_key = f"{org_id}/{project_id}"
            else:
                # If not in compliance system, use project_id directly (format: n/a/project_id)
                account_key = f"n/a/{project_id}"
            
            accounts.append(account_key)
        
        # Log found integrations
        if found_integrations:
            Logger.info(f"Found {len(found_integrations)} GcpCfg integration(s) with state.ok == True:")
            for project_id, name in found_integrations:
                Logger.info(f"  {name} (project: {project_id})")
        
        # Log skipped accounts
        if skipped_accounts:
            Logger.warning(f"Found {len(skipped_accounts)} project(s) with integration errors:")
            for project_id, name in skipped_accounts:
                Logger.warning(f"  {name} (project: {project_id}): GcpCfg integration error (state.ok != True)")
            Logger.warning("These projects have integration issues and have been skipped.")
        
        if not accounts:
            Logger.warning("No GCP projects found with GcpCfg integrations (state.ok == True)")
        
        return sorted(set(accounts))


def get_accounts(
    cloud_type: CloudType,
    env: Dict[str, str],
    verbose: bool,
    use_cache: bool,
    cache_dir: str
) -> List[str]:
    """Get accounts for the specified cloud type."""
    fetcher_classes = {
        CloudType.AWS: AWSAccountFetcher,
        CloudType.AZURE: AzureAccountFetcher,
        CloudType.GCP: GCPAccountFetcher,
    }
    
    fetcher_class = fetcher_classes.get(cloud_type)
    if not fetcher_class:
        Logger.error(f"Unknown cloud type: {cloud_type}")
        sys.exit(1)

    fetcher = fetcher_class(cloud_type, env, verbose, cache_dir)
    return fetcher.get_accounts(use_cache)


# ============================================================================
# Report Fetching
# ============================================================================

def get_report_for_account(
    cloud_type: CloudType,
    report_name: str,
    account: str,
    output_file: str,
    env: Dict[str, str],
    verbose: bool
) -> bool:
    """Get report for a specific account via /api/v2/Reports."""
    Logger.verbose(f"Fetching report '{report_name}' for {cloud_type.value} account: {account}", verbose)

    cmd = _build_report_command(cloud_type, report_name, account)
    output, _ = make_api_call(cmd, env, verbose)

    if not output:
        Logger.warning(f"No report data returned for account: {account}")
        Logger.warning("This account may be disabled or have no compliance data available")
        return False

    # Parse and unwrap the API response
    try:
        api_response = json.loads(output)
    except (json.JSONDecodeError, ValueError):
        Logger.warning(f"Invalid JSON response for account: {account}")
        Logger.verbose(f"Response was: {output}", verbose)
        return False

    report_data = _unwrap_api_response(api_response, account, verbose)
    if report_data is None:
        return False

    # Log violation object structure for debugging
    if verbose:
        recs = report_data.get('recommendations', [])
        for r in recs:
            violations = r.get('VIOLATIONS', [])
            if violations:
                Logger.verbose(f"Violation object keys: {list(violations[0].keys())}", verbose)
                Logger.verbose(f"Sample violation: {json.dumps(violations[0], indent=2)}", verbose)
                break

    # Save the unwrapped report data
    with open(output_file, 'w') as f:
        json.dump(report_data, f, indent=2)

    Logger.verbose(f"Saved report for account '{account}' to: {output_file}", verbose)
    return True


def _build_report_command(cloud_type: CloudType, report_name: str, account: str) -> List[str]:
    """Build Lacework API command for getting a report via /api/v2/Reports."""
    encoded_name = urllib.parse.quote(report_name, safe='')
    params = f"format=json&reportName={encoded_name}"

    if cloud_type == CloudType.AWS:
        params += f"&primaryQueryId={account}"
    elif cloud_type == CloudType.AZURE:
        tenant_id, subscription_id = account.split('/', 1)
        params += f"&primaryQueryId={tenant_id}&secondaryQueryId={subscription_id}"
    elif cloud_type == CloudType.GCP:
        org_id, project_id = account.split('/', 1)
        params += f"&primaryQueryId={org_id}&secondaryQueryId={project_id}"
    else:
        Logger.error(f"Unknown cloud type: {cloud_type}")
        sys.exit(1)

    return ['lacework', 'api', 'get', f'api/v2/Reports?{params}', '--json', '--noninteractive']


def _unwrap_api_response(api_response, account: str, verbose: bool) -> Optional[Dict]:
    """
    Unwrap /api/v2/Reports response to match the format expected by downstream consumers.

    API returns: {"data": [{reportType, reportTitle, recommendations, summary, ...}]}
    Old CLI returned: {reportType, reportTitle, recommendations, summary, accountId, ...}

    Extracts data[0] and ensures accountId is populated.
    """
    if isinstance(api_response, dict) and 'data' in api_response:
        data_list = api_response['data']
        if not data_list:
            Logger.warning(f"Empty report data for account: {account}")
            return None
        report_data = data_list[0]
    elif isinstance(api_response, dict):
        report_data = api_response
    else:
        Logger.warning(f"Unexpected response format for account: {account}")
        Logger.verbose(f"Response type: {type(api_response)}", verbose)
        return None

    # Inject accountId if missing (old CLI included it, API may not)
    if not report_data.get('accountId'):
        report_data['accountId'] = account

    # Fallback reportTime if missing
    if not report_data.get('reportTime'):
        report_data['reportTime'] = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')

    return report_data


# ============================================================================
# LQL-Based Report Generation (default)
# ============================================================================

def _fetch_report_definition(
    report_name: str,
    env: Dict[str, str],
    verbose: bool
) -> Optional[Dict]:
    """Fetch a report definition by name from GET /api/v2/ReportDefinitions."""
    cmd = ['lacework', 'api', 'get', 'api/v2/ReportDefinitions', '--json', '--noninteractive']
    output, _ = make_api_call(cmd, env, verbose)
    if not output:
        return None

    try:
        response = json.loads(output)
        defs = response.get('data', response) if isinstance(response, dict) else response
        for d in defs:
            if d.get('reportName') == report_name:
                return d
    except (json.JSONDecodeError, ValueError) as e:
        Logger.warning(f"Failed to parse ReportDefinitions response: {e}")
    return None


def _fetch_all_policies(
    env: Dict[str, str],
    verbose: bool
) -> Dict[str, Dict]:
    """Fetch all policies and build a {policyId: metadata} lookup.

    Severity is converted from string to numeric (1-5).
    """
    cmd = ['lacework', 'api', 'get', 'api/v2/Policies', '--json', '--noninteractive']
    output, _ = make_api_call(cmd, env, verbose)
    if not output:
        return {}

    try:
        response = json.loads(output)
        policies = response.get('data', response) if isinstance(response, dict) else response
    except (json.JSONDecodeError, ValueError) as e:
        Logger.warning(f"Failed to parse Policies response: {e}")
        return {}

    lookup = {}
    for p in policies:
        pid = p.get('policyId', '')
        if not pid:
            continue
        sev_str = p.get('severity', 'info')
        # Tags come as list of "key:value" strings — convert to dict
        raw_tags = p.get('tags', [])
        tags_dict = {}
        if isinstance(raw_tags, list):
            for t in raw_tags:
                if ':' in t:
                    k, v = t.split(':', 1)
                    tags_dict[k] = v
        elif isinstance(raw_tags, dict):
            tags_dict = raw_tags
        lookup[pid] = {
            'title': p.get('title', ''),
            'severity': CONFIG.SEVERITY_STRING_TO_NUM.get(sev_str.lower(), 5),
            'queryId': p.get('queryId'),
            'infoLink': p.get('infoLink', ''),
            'tags': tags_dict,
            'description': p.get('description', ''),
            'remediation': p.get('remediation', ''),
        }
    return lookup


def _execute_policy_query(
    query_id: str,
    env: Dict[str, str],
    verbose: bool
) -> List[Dict]:
    """Execute a policy's LQL query via POST /api/v2/Queries/{queryId}/execute.

    Returns list of violation rows. Each row typically contains ACCOUNT_ID,
    RESOURCE_KEY, RESOURCE_REGION, COMPLIANCE_FAILURE_REASON, etc.
    """
    now = datetime.datetime.now(datetime.timezone.utc)
    end_time = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    start_time = (now - datetime.timedelta(days=7)).strftime('%Y-%m-%dT%H:%M:%SZ')

    payload = json.dumps({
        "arguments": [
            {"name": "StartTimeRange", "value": start_time},
            {"name": "EndTimeRange", "value": end_time}
        ]
    })

    # Validate query_id to prevent path traversal
    if not re.match(r'^[A-Za-z0-9_-]+$', query_id):
        Logger.warning(f"Invalid query_id format: {query_id} — skipping")
        return []

    cmd = [
        'lacework', 'api', 'post', f'api/v2/Queries/{query_id}/execute',
        '-d', payload, '--json', '--noninteractive'
    ]

    output, was_rate_limited = make_api_call(cmd, env, verbose)
    if was_rate_limited or not output:
        return []

    try:
        response = json.loads(output)
        data = response.get('data', []) if isinstance(response, dict) else []
        return data
    except (json.JSONDecodeError, ValueError) as e:
        Logger.warning(f"Failed to parse query response for {query_id}: {e}")
        return []


def _build_recommendation_from_query(
    policy_id: str,
    policy_meta: Dict,
    violations: List[Dict],
    account_id: str,
    category: str
) -> Dict:
    """Build a recommendation dict matching the /api/v2/Reports format."""
    if violations:
        status = 'NonCompliant'
    elif policy_meta.get('queryId'):
        status = 'Compliant'
    else:
        status = 'CouldNotAssess'

    mapped_violations = []
    for v in violations:
        mapped_violations.append({
            'resource': v.get('RESOURCE_KEY', ''),
            'region': v.get('RESOURCE_REGION', ''),
            'reasonDescription': v.get('COMPLIANCE_FAILURE_REASON', ''),
        })

    return {
        'REC_ID': policy_id,
        'TITLE': policy_meta.get('title', ''),
        'SEVERITY': policy_meta.get('severity', 5),
        'CATEGORY': category,
        'SERVICE': policy_meta.get('tags', {}).get('subdomain', ''),
        'STATUS': status,
        'NUM_VIOLATIONS': len(violations),
        'ASSESSED_RESOURCE_COUNT': 0,
        'RESOURCE_COUNT': 0,
        'SUPPRESSED_RESOURCE_COUNT': 0,
        'INFO_LINK': policy_meta.get('infoLink', ''),
        'REMEDIATION': policy_meta.get('remediation', ''),
        'ACCOUNT_ID': account_id,
        'VIOLATIONS': mapped_violations,
    }


def _fetch_reports_via_lql(
    report_name: str,
    cloud_type: CloudType,
    accounts: List[str],
    report_output_dir: str,
    env: Dict[str, str],
    verbose: bool,
    test_mode: bool = False
) -> Optional[Tuple[int, int]]:
    """
    Fetch compliance report via individual LQL policy queries.

    Fallback for when /api/v2/Reports returns 500 errors. Executes each
    policy's underlying query, groups violations by account, and writes
    per-account JSON files in the same format as get_report_for_account().

    Returns (success_count, failure_count), or None if the report definition
    is not found (caller should fall back to Reports API).
    """
    # 1. Fetch report definition
    Logger.verbose(f"Fetching report definition for '{report_name}'...", verbose)
    report_def = _fetch_report_definition(report_name, env, verbose)
    if not report_def:
        return None  # Signal caller to fall back to Reports API

    sections = report_def.get('reportDefinition', {}).get('sections', [])
    if not sections:
        Logger.error(f"Report '{report_name}' has no sections")
        sys.exit(1)

    # 2. Fetch all policies
    Logger.info("Fetching policy metadata...")
    all_policies = _fetch_all_policies(env, verbose)
    if not all_policies:
        Logger.error("Failed to fetch policies")
        sys.exit(1)
    Logger.info(f"Loaded metadata for {len(all_policies)} policies")

    # 3. Extract policy IDs from sections (deduplicate across sections)
    policy_section_map: Dict[str, str] = {}  # {policy_id: section_title}
    all_policy_ids: List[str] = []
    seen_pids: set = set()
    duplicates_skipped = 0
    for section in sections:
        section_title = section.get('title', section.get('category', ''))
        for pid in section.get('policies', []):
            if pid in seen_pids:
                duplicates_skipped += 1
                Logger.verbose(
                    f"Skipping duplicate policy {pid} "
                    f"(already in '{policy_section_map[pid]}', "
                    f"also in '{section_title}')", verbose)
                continue
            seen_pids.add(pid)
            policy_section_map[pid] = section_title
            all_policy_ids.append(pid)

    Logger.info(
        f"Report has {len(sections)} sections, "
        f"{len(all_policy_ids)} unique policies"
        + (f" ({duplicates_skipped} cross-section duplicates skipped)"
           if duplicates_skipped else "")
    )

    # Test mode: limit to first 3 policies per section
    if test_mode:
        limited_ids = []
        section_counts: Dict[str, int] = {}
        for pid in all_policy_ids:
            cat = policy_section_map[pid]
            section_counts[cat] = section_counts.get(cat, 0) + 1
            if section_counts[cat] <= 3:
                limited_ids.append(pid)
        all_policy_ids = limited_ids
        Logger.info(
            f"Test mode: limiting to {len(all_policy_ids)} policies "
            f"(3 per section)"
        )

    # 4. Execute each policy query
    # {account_id: {policy_id: [violations]}}
    account_violations: Dict[str, Dict[str, List[Dict]]] = {}
    null_query_policies: List[str] = []
    policy_count = len(all_policy_ids)

    def _extract_account_key(violation: Dict, ct: CloudType) -> str:
        """Extract account key from a violation row.

        AWS queries return ACCOUNT_ID. Azure returns TENANT_ID/SUBSCRIPTION_ID.
        GCP returns PROJECT_ID (with org handled at account discovery level).
        """
        if ct == CloudType.AZURE:
            tenant = violation.get('TENANT_ID', '')
            sub = violation.get('SUBSCRIPTION_ID', '')
            if tenant and sub:
                return f"{tenant}/{sub}"
            return ''
        if ct == CloudType.GCP:
            return violation.get('PROJECT_ID', violation.get('ACCOUNT_ID', ''))
        return violation.get('ACCOUNT_ID', '')

    for i, policy_id in enumerate(all_policy_ids):
        policy_meta = all_policies.get(policy_id)
        if not policy_meta:
            Logger.warning(f"Policy {policy_id} not found — skipping")
            continue

        query_id = policy_meta.get('queryId')
        if not query_id:
            null_query_policies.append(policy_id)
            Logger.verbose(
                f"[{i + 1}/{policy_count}] {policy_id}: "
                f"no queryId (manual assessment)", verbose
            )
            continue

        Logger.info(f"[{i + 1}/{policy_count}] Querying: {policy_id}")

        violations = _execute_policy_query(query_id, env, verbose)

        if violations:
            acct_keys = set(_extract_account_key(v, cloud_type) for v in violations)
            acct_keys.discard('')
            Logger.verbose(
                f"  {len(violations)} violations across "
                f"{len(acct_keys)} accounts",
                verbose
            )
        else:
            Logger.verbose(f"  0 violations", verbose)

        # Group violations by account key
        for v in violations:
            acct = _extract_account_key(v, cloud_type)
            if not acct:
                continue
            if acct not in account_violations:
                account_violations[acct] = {}
            if policy_id not in account_violations[acct]:
                account_violations[acct][policy_id] = []
            account_violations[acct][policy_id].append(v)

        # Delay between queries
        if i < policy_count - 1:
            time.sleep(CONFIG.REQUEST_DELAY)

    # Log warnings for special cases
    if null_query_policies:
        Logger.info(
            f"{len(null_query_policies)} policies have no queryId "
            f"(manual assessments — shown as CouldNotAssess)"
        )

    # 5. Build per-account JSON files
    report_time = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')

    # All accounts: discovered + those found in query results
    all_account_ids = set(accounts) | set(account_violations.keys())
    Logger.info(
        f"Building reports for {len(all_account_ids)} accounts "
        f"({len(account_violations)} with violations)..."
    )

    success_count = 0
    for account_id in sorted(all_account_ids):
        recommendations = []
        num_compliant = 0
        num_non_compliant = 0
        severity_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        total_violated_resources = 0

        acct_viols = account_violations.get(account_id, {})

        for policy_id in all_policy_ids:
            policy_meta = all_policies.get(policy_id, {})
            category = policy_section_map.get(policy_id, '')
            violations_for_policy = acct_viols.get(policy_id, [])

            rec = _build_recommendation_from_query(
                policy_id, policy_meta, violations_for_policy,
                account_id, category
            )
            recommendations.append(rec)

            if rec['STATUS'] == 'NonCompliant':
                num_non_compliant += 1
                sev = policy_meta.get('severity', 5)
                severity_counts[sev] = severity_counts.get(sev, 0) + 1
                total_violated_resources += rec['NUM_VIOLATIONS']
            elif rec['STATUS'] == 'Compliant':
                num_compliant += 1

        summary = {
            'NUM_RECOMMENDATIONS': len(recommendations),
            'NUM_COMPLIANT': num_compliant,
            'NUM_NOT_COMPLIANT': num_non_compliant,
            'NUM_SEVERITY_1_NON_COMPLIANCE': severity_counts.get(1, 0),
            'NUM_SEVERITY_2_NON_COMPLIANCE': severity_counts.get(2, 0),
            'NUM_SEVERITY_3_NON_COMPLIANCE': severity_counts.get(3, 0),
            'NUM_SEVERITY_4_NON_COMPLIANCE': severity_counts.get(4, 0),
            'NUM_SEVERITY_5_NON_COMPLIANCE': severity_counts.get(5, 0),
            'NUM_SUPPRESSED': 0,
            'ASSESSED_RESOURCE_COUNT': 0,
            'VIOLATED_RESOURCE_COUNT': total_violated_resources,
            'SUPPRESSED_RESOURCE_COUNT': 0,
        }

        report_data = {
            'reportTitle': report_name,
            'reportType': 'COMPLIANCE',
            'reportTime': report_time,
            'accountId': account_id,
            'recommendations': recommendations,
            'summary': [summary]
        }

        safe_name = account_id.replace('/', '_').replace(':', '_')
        output_file = os.path.join(report_output_dir, f"{safe_name}.json")
        with open(output_file, 'w') as f:
            json.dump(report_data, f, indent=2)

        success_count += 1

    Logger.info(
        f"LQL fallback complete: {success_count} account reports generated"
    )
    return success_count, 0


# ============================================================================
# Excel Generation
# ============================================================================

def create_excel_from_report(data: Dict, output_file: str, include_compliant: bool = False, tags_lookup: Dict = None) -> None:
    """Create Excel spreadsheet from report data."""
    if not HAS_OPENPYXL:
        Logger.error("openpyxl is required for Excel output. Install it with: pip3 install openpyxl")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _create_summary_sheet(wb, data)
    _create_recommendations_sheet(wb, data, include_compliant, tags_lookup)

    wb.save(output_file)


def _create_summary_sheet(wb: Workbook, data: Dict) -> None:
    """Create Summary sheet with overview and non-compliant policy breakdown."""
    ws = wb.create_sheet("Summary", 0)
    summary = data['summary'][0]
    recommendations = data.get('recommendations', [])

    # Calculate overview stats
    total_policies = summary.get('NUM_RECOMMENDATIONS', 0)
    non_compliant_count = summary.get('NUM_NOT_COMPLIANT', 0)
    compliant_count = summary.get('NUM_COMPLIANT', 0)
    could_not_assess_count = sum(
        1 for rec in recommendations
        if rec.get('STATUS', '').upper() == 'COULDNOTASSESS'
    )
    non_compliant_pct = (
        (non_compliant_count / total_policies * 100) if total_policies > 0 else 0
    )

    # Unique accounts
    all_accounts = set(rec.get('ACCOUNT_ID', '') for rec in recommendations)
    all_accounts.discard('')

    # Header
    ws['A1'] = data['reportTitle']
    ws['A1'].font = Font(bold=True, size=16, color="1A1A1A")
    ws['A2'] = f"Report Time: {data['reportTime']}"
    ws['A2'].font = Font(size=10, color="333333")

    # Overview metrics
    row = 4
    _add_section_header(ws, row, 'Overview', num_cols=3)
    row += 1
    _add_metric_row(ws, row, 'Total Accounts', len(all_accounts))
    row += 1
    _add_metric_row(ws, row, 'Total Policies', total_policies)
    row += 1
    _add_metric_row(ws, row, 'Non-Compliant', f"{non_compliant_count} ({non_compliant_pct:.1f}%)")
    row += 1
    _add_metric_row(ws, row, 'Compliant', compliant_count)
    row += 1
    _add_metric_row(ws, row, 'Could Not Assess', could_not_assess_count)

    # Build per-policy stats: {policy_title: {severity, non_compliant_accounts, compliant_accounts}}
    policy_stats: Dict[str, Dict] = {}
    for rec in recommendations:
        title = rec.get('TITLE', '')
        if not title:
            continue
        if title not in policy_stats:
            policy_stats[title] = {
                'severity': rec.get('SEVERITY', 5),
                'rec_id': rec.get('REC_ID', ''),
                'non_compliant': set(),
                'compliant': set(),
            }
        acct = rec.get('ACCOUNT_ID', '')
        status = rec.get('STATUS', '').upper()
        if status == 'NONCOMPLIANT':
            policy_stats[title]['non_compliant'].add(acct)
        elif status == 'COMPLIANT':
            policy_stats[title]['compliant'].add(acct)

    # Filter to non-compliant policies, sort by severity then non-compliant count
    nc_policies = [
        (title, stats) for title, stats in policy_stats.items()
        if stats['non_compliant']
    ]
    nc_policies.sort(key=lambda x: (x[1]['severity'], -len(x[1]['non_compliant'])))

    # Non-Compliant Policies table
    row += 2
    _add_section_header(ws, row, 'Non-Compliant Policies', num_cols=3)
    row += 1

    # Table header
    table_headers = ['Policy', 'Severity', 'Non-Compliant Accounts', 'Compliant Accounts']
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    for col, hdr in enumerate(table_headers, 1):
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    # Table rows
    for title, stats in nc_policies:
        sev_label = CONFIG.SEVERITY_MAP.get(stats['severity'], 'Unknown')
        nc_count = len(stats['non_compliant'])
        c_count = len(stats['compliant'])

        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=sev_label)
        ws.cell(row=row, column=3, value=nc_count)
        ws.cell(row=row, column=4, value=c_count)

        # Color severity
        sev_colors = {1: 'DC2626', 2: 'EA580C', 3: 'D97706', 4: '2563EB', 5: '6B7280'}
        sev_color = sev_colors.get(stats['severity'], '6B7280')
        ws.cell(row=row, column=2).font = Font(size=10, color=sev_color)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25


def _add_section_header(ws, row: int, title: str, num_cols: int = 2) -> None:
    """Add a section header with vibrant styling."""
    header_cell = ws[f'A{row}']
    header_cell.value = title
    header_cell.font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_cell.fill = header_fill
    end_col = get_column_letter(max(num_cols, 2))
    ws.merge_cells(f'A{row}:{end_col}{row}')
    # Apply fill to all merged cells
    for col in range(2, max(num_cols, 2) + 1):
        ws.cell(row=row, column=col).fill = header_fill
    ws.row_dimensions[row].height = 24
    header_cell.alignment = Alignment(horizontal='left', vertical='center')


def _add_metric_row(ws, row: int, label: str, value, right_align: bool = False) -> None:
    """Add a metric row with vibrant styling."""
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = Font(size=10, color="1A1A1A")
    ws[f'B{row}'] = value
    ws[f'B{row}'].font = Font(size=10, color="000000")
    if right_align:
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 18


def _collect_violation_urns(recommendations: List[Dict]) -> List[str]:
    """Collect unique resource URNs from all violations across recommendations.

    Matches AWS ARNs (arn:...) and Azure resource-level paths
    (/subscriptions/.../resourceGroups/...). Skips bare subscription IDs,
    location paths, and other non-resource identifiers.
    """
    urns = set()
    for rec in recommendations:
        for v in rec.get('VIOLATIONS', []):
            resource = v.get('resource', '')
            if not resource:
                continue
            # AWS ARNs
            if resource.startswith('arn:'):
                urns.add(resource)
            # Azure resource-level paths (skip subscription-only and location paths)
            elif '/resourcegroups/' in resource.lower():
                urns.add(resource)
    return list(urns)


def _get_azure_parent_urn(urn: str) -> Optional[str]:
    """Strip sub-resource path segments to get the parent Azure resource URN.

    E.g. .../storageaccounts/foo/blobservices/default/containers/bar
      -> .../storageaccounts/foo
    """
    lower = urn.lower()
    idx = lower.find('/providers/')
    if idx < 0:
        return None
    provider_path = urn[idx + len('/providers/'):]
    parts = provider_path.split('/')
    if len(parts) <= 3:
        return None  # Already at parent level
    return urn[:idx] + '/providers/' + '/'.join(parts[:3])


def _fetch_azure_tags_via_lql(
    urns: List[str],
    env: Dict[str, str],
    verbose: bool
) -> Dict[str, Dict]:
    """
    Fetch Azure resource tags via the LW_CFG_AZURE_ALL LQL datasource.

    The Inventory API (/api/v2/Inventory/search) returns empty for Azure,
    so we query LQL directly via POST /api/v2/Queries/execute, partitioned
    by subscription ID to stay under the 5000-row query cap.

    Returns {urn: {tag_key: tag_value}} for resources with non-empty tags.
    Includes parent-URN fallback for sub-resources (e.g. blob containers
    inherit tags from their parent storage account).
    """
    # Extract unique subscription IDs from violation URNs
    sub_ids = set()
    for urn in urns:
        match = re.search(r'/subscriptions/([^/]+)/', urn, re.IGNORECASE)
        if match:
            sub_ids.add(match.group(1).lower())

    if not sub_ids:
        Logger.warning("No subscription IDs found in Azure violation URNs")
        return {}

    Logger.info(
        f"Fetching Azure tags via LQL (LW_CFG_AZURE_ALL) "
        f"for {len(sub_ids)} subscriptions..."
    )

    # Time range: last 7 days
    now = datetime.datetime.now(datetime.timezone.utc)
    end_time = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    start_time = (now - datetime.timedelta(days=7)).strftime('%Y-%m-%dT%H:%M:%SZ')

    # Build tag lookup keyed by lowercase URN
    raw_lookup: Dict[str, Dict] = {}  # lowercase_urn -> tags
    total_tagged = 0

    for i, sub_id in enumerate(sorted(sub_ids)):
        query_text = (
            f"{{ source {{ LW_CFG_AZURE_ALL r }} "
            f"filter {{ r.URN LIKE '/subscriptions/{sub_id}/%' "
            f"AND r.RESOURCE_TAGS <> '{{}}' "
            f"AND r.RESOURCE_TAGS is not null }} "
            f"return distinct {{ r.URN, r.RESOURCE_TAGS }} }}"
        )

        payload = json.dumps({
            "query": {"queryText": query_text},
            "arguments": [
                {"name": "StartTimeRange", "value": start_time},
                {"name": "EndTimeRange", "value": end_time}
            ]
        })

        cmd = [
            'lacework', 'api', 'post', 'api/v2/Queries/execute',
            '-d', payload, '--json', '--noninteractive'
        ]

        try:
            output, was_rate_limited = make_api_call(cmd, env, verbose)

            if was_rate_limited:
                Logger.warning(
                    f"Rate-limited on subscription {i + 1}/{len(sub_ids)} "
                    f"({sub_id}) — skipping"
                )
                continue

            if not output:
                Logger.verbose(
                    f"  Subscription {i + 1}/{len(sub_ids)} ({sub_id}): "
                    f"empty response", verbose
                )
                continue

            response = json.loads(output)
            data = response.get('data', []) if isinstance(response, dict) else []

            for item in data:
                urn_val = item.get('URN', '')
                tags = item.get('RESOURCE_TAGS')
                if urn_val and tags and isinstance(tags, dict):
                    urn_lower = urn_val.lower()
                    if urn_lower not in raw_lookup:
                        raw_lookup[urn_lower] = tags

            Logger.info(
                f"  Subscription {i + 1}/{len(sub_ids)} ({sub_id}): "
                f"{len(data)} tagged resources"
            )
            total_tagged += len(data)

        except (json.JSONDecodeError, ValueError) as e:
            Logger.warning(
                f"Failed to parse LQL response for subscription "
                f"{sub_id}: {e}"
            )
        except Exception as e:
            Logger.warning(
                f"Error fetching LQL tags for subscription {sub_id}: {e}"
            )

        # Delay between subscription queries
        if i < len(sub_ids) - 1:
            time.sleep(CONFIG.TAG_FETCH_DELAY)

    Logger.info(
        f"Azure tag lookup: {len(raw_lookup)} unique tagged resources "
        f"across {len(sub_ids)} subscriptions"
    )

    # Match violation URNs to tags (case-insensitive, with parent fallback)
    tags_lookup: Dict[str, Dict] = {}
    parent_matches = 0

    for urn in urns:
        urn_lower = urn.lower()
        if urn_lower in raw_lookup:
            tags_lookup[urn] = raw_lookup[urn_lower]
        else:
            parent = _get_azure_parent_urn(urn)
            if parent and parent.lower() in raw_lookup:
                tags_lookup[urn] = raw_lookup[parent.lower()]
                parent_matches += 1

    Logger.info(
        f"Matched tags for {len(tags_lookup)}/{len(urns)} violation resources"
        + (f" ({parent_matches} via parent fallback)" if parent_matches else "")
    )

    return tags_lookup


def _fetch_inventory_tags(
    urns: List[str],
    cloud_type: CloudType,
    env: Dict[str, str],
    verbose: bool
) -> Dict[str, Dict]:
    """
    Batch-fetch resource tags from the Inventory API.

    Calls POST /api/v2/Inventory/search with URN filters in batches of 100.
    Returns {urn: {tag_key: tag_value}} for resources with non-empty tags.
    On error, warns and returns partial results (never fails the report).

    Azure uses a separate LQL-based path since the Inventory API returns
    empty results for Azure resources.
    """
    # Azure: route to LQL-based tag fetching
    if cloud_type == CloudType.AZURE:
        return _fetch_azure_tags_via_lql(urns, env, verbose)

    csp_map = {
        CloudType.AWS: 'AWS',
        CloudType.GCP: 'GCP',
    }
    csp = csp_map.get(cloud_type)
    if not csp:
        Logger.warning(f"Unknown cloud type for inventory lookup: {cloud_type}")
        return {}

    BATCH_SIZE = 100
    MAX_REQUEUES = 2
    RATE_LIMIT_COOLDOWN = 60  # seconds to wait after rate-limit exhaustion
    tags_lookup: Dict[str, Dict] = {}

    # Build work queue: list of (batch_urns, requeue_count)
    batches: List[Tuple[List[str], int]] = []
    for i in range(0, len(urns), BATCH_SIZE):
        batches.append((urns[i:i + BATCH_SIZE], 0))

    total_batches = len(batches)
    fetched_resources = 0
    batch_index = 0

    while batch_index < len(batches):
        batch, requeue_count = batches[batch_index]
        Logger.info(
            f"Fetching tags for {fetched_resources}/{len(urns)} resources "
            f"(batch {batch_index + 1}/{len(batches)})"
        )
        Logger.verbose(f"  Batch has {len(batch)} URNs (requeue #{requeue_count})", verbose)

        payload = json.dumps({
            "csp": csp,
            "filters": [
                {"field": "urn", "expression": "in", "values": batch}
            ],
            "returns": ["urn", "resourceTags"]
        })

        cmd = [
            'lacework', 'api', 'post', 'api/v2/Inventory/search',
            '-d', payload, '--json', '--noninteractive'
        ]

        try:
            output, was_rate_limited = make_api_call(cmd, env, verbose)

            if was_rate_limited:
                if requeue_count < MAX_REQUEUES:
                    Logger.warning(
                        f"Rate-limited on batch {batch_index + 1} — "
                        f"cooling down {RATE_LIMIT_COOLDOWN}s then retrying "
                        f"(attempt {requeue_count + 1}/{MAX_REQUEUES})"
                    )
                    time.sleep(RATE_LIMIT_COOLDOWN)
                    batches[batch_index] = (batch, requeue_count + 1)
                    continue  # retry same batch_index without advancing
                else:
                    Logger.warning(
                        f"Rate-limited on batch {batch_index + 1} after "
                        f"{MAX_REQUEUES} re-queues — skipping batch"
                    )
                    batch_index += 1
                    continue

            if not output:
                Logger.warning(f"Empty response for inventory batch {batch_index + 1}")
                batch_index += 1
                continue

            response = json.loads(output)
            data = response.get('data', []) if isinstance(response, dict) else []

            for item in data:
                urn = item.get('urn', '')
                resource_tags = item.get('resourceTags', {})
                # First non-empty tags per URN wins (multiple rows possible per URN)
                if urn and resource_tags and urn not in tags_lookup:
                    tags_lookup[urn] = resource_tags

            fetched_resources += len(batch)

        except (json.JSONDecodeError, ValueError) as e:
            Logger.warning(f"Failed to parse inventory response for batch {batch_index + 1}: {e}")
        except Exception as e:
            Logger.warning(f"Error fetching inventory batch {batch_index + 1}: {e}")

        batch_index += 1

        # Delay between batches
        if batch_index < len(batches):
            time.sleep(CONFIG.TAG_FETCH_DELAY)

    Logger.info(f"Tag lookup complete: {len(tags_lookup)}/{len(urns)} resources with tags")
    return tags_lookup


# ============================================================================
# Violation History Tracking
# ============================================================================

def _get_history_path(report_name: str) -> str:
    """Get the CSV history file path for a report."""
    safe_name = report_name.replace('/', '_').replace('\\', '_').replace('..', '_')
    return os.path.join(CONFIG.HISTORY_DIR, f"{safe_name}.csv")


def _load_violation_history(report_name: str) -> Dict[tuple, str]:
    """Load violation history from CSV. Returns {(policy_id, account_id, resource): first_seen_date}."""
    path = _get_history_path(report_name)
    history = {}
    if not os.path.exists(path):
        return history
    with open(path, 'r', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row.get('policy_id', ''), row.get('account_id', ''), row.get('resource', ''))
            history[key] = row.get('first_seen', '')
    return history


def _save_violation_history(report_name: str, history: Dict[tuple, str]) -> None:
    """Save violation history to CSV."""
    os.makedirs(CONFIG.HISTORY_DIR, exist_ok=True)
    path = _get_history_path(report_name)
    with open(path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['policy_id', 'account_id', 'resource', 'first_seen'])
        writer.writeheader()
        for (policy_id, account_id, resource), first_seen in sorted(history.items()):
            writer.writerow({
                'policy_id': policy_id,
                'account_id': account_id,
                'resource': resource,
                'first_seen': first_seen,
            })


def _enrich_with_first_seen(
    recommendations: List[Dict],
    report_name: str,
    verbose: bool
) -> None:
    """Enrich violations with first-seen dates from history.

    Loads existing history, sets first_seen on each violation, records new
    violations with today's date, and saves updated history.
    """
    history = _load_violation_history(report_name)
    today = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d')
    new_count = 0

    for rec in recommendations:
        rec_id = rec.get('REC_ID', '')
        account_id = rec.get('ACCOUNT_ID', '')
        for v in rec.get('VIOLATIONS', []):
            resource = v.get('resource', '')
            key = (rec_id, account_id, resource)
            if key not in history:
                history[key] = today
                new_count += 1
            v['first_seen'] = history[key]

    _save_violation_history(report_name, history)

    total = len(history)
    Logger.info(f"Violation history: {total} tracked, {new_count} new today")
    Logger.verbose(f"History file: {_get_history_path(report_name)}", verbose)


def _expand_recommendations_to_rows(recommendations: List[Dict], include_compliant: bool = False, tags_lookup: Dict = None) -> List[Dict]:
    """
    Expand recommendations so each violation gets its own row.

    For non-compliant recommendations with violations: one row per violation,
    carrying the policy metadata plus the individual resource and tags.
    For compliant recommendations (when include_compliant=True): one row with
    empty Resource/Tags.
    For non-compliant with no violations: one row with empty Resource/Tags.

    Deduplicates by (REC_ID, ACCOUNT_ID, resource) to handle policies that
    appear in multiple report sections with identical violations.
    """
    rows = []
    seen_keys: set = set()  # (REC_ID, ACCOUNT_ID, resource) for dedup
    duplicates_skipped = 0
    for rec in recommendations:
        status = rec.get('STATUS', '')
        violations = rec.get('VIOLATIONS', [])
        is_non_compliant = status.lower() == 'noncompliant'

        if not is_non_compliant and not include_compliant:
            continue

        rec_id = rec.get('REC_ID', '')
        account_id = rec.get('ACCOUNT_ID', '')

        # Base row data shared across all expanded rows for this recommendation
        base = {
            'CATEGORY': rec.get('CATEGORY', ''),
            'TITLE': rec.get('TITLE', ''),
            'REC_ID': rec_id,
            'SEVERITY': rec.get('SEVERITY', ''),
            'ACCOUNT_ID': account_id,
            'ACCOUNT_ALIAS': rec.get('ACCOUNT_ALIAS', ''),
            'STATUS': status,
            'REMEDIATION': rec.get('REMEDIATION', ''),
        }

        if violations:
            for v in violations:
                resource = v.get('resource', '')
                dedup_key = (rec_id, account_id, resource)
                if dedup_key in seen_keys:
                    duplicates_skipped += 1
                    continue
                seen_keys.add(dedup_key)
                row = dict(base)
                row['RESOURCE'] = resource
                row['FIRST_SEEN'] = v.get('first_seen', '')
                # Prefer tags from inventory lookup, fall back to inline tags
                tags = (tags_lookup.get(resource, {}) if tags_lookup else {}) or v.get('resourceTags') or v.get('tags') or {}
                row['TAGS'] = json.dumps(tags) if tags else ''
                rows.append(row)
        else:
            dedup_key = (rec_id, account_id, '')
            if dedup_key in seen_keys:
                duplicates_skipped += 1
                continue
            seen_keys.add(dedup_key)
            row = dict(base)
            row['RESOURCE'] = ''
            row['FIRST_SEEN'] = ''
            row['TAGS'] = ''
            rows.append(row)

    if duplicates_skipped:
        Logger.info(
            f"Deduplicated {duplicates_skipped} cross-section duplicate "
            f"violation(s)")

    return rows


def _create_recommendations_sheet(wb: Workbook, data: Dict, include_compliant: bool = False, tags_lookup: Dict = None) -> None:
    """Create Recommendations sheet in workbook."""
    ws = wb.create_sheet("Recommendations", 1)
    recommendations = data.get('recommendations', [])

    if not recommendations:
        return

    # Expand violations to individual rows
    expanded_rows = _expand_recommendations_to_rows(recommendations, include_compliant, tags_lookup)

    if not expanded_rows:
        return

    # Sort expanded rows
    expanded_rows = sorted(expanded_rows, key=_recommendation_sort_key)

    # Write headers
    _write_excel_headers(ws, CONFIG.EXCEL_HEADERS)

    # Write data rows
    for row_num, rec in enumerate(expanded_rows, 2):
        _write_recommendation_row(ws, row_num, rec)

    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-adjust column widths
    _adjust_column_widths(ws, len(CONFIG.EXCEL_HEADERS))


def _recommendation_sort_key(rec: Dict) -> Tuple:
    """Generate sort key for recommendations."""
    category = rec.get('CATEGORY', '')
    service = rec.get('SERVICE', '')
    severity_num = rec.get('SEVERITY', 99)
    severity_label = CONFIG.SEVERITY_MAP.get(severity_num, 'Unknown')
    rec_id = rec.get('REC_ID', '')
    account_id = rec.get('ACCOUNT_ID', '')
    resource = rec.get('RESOURCE', '')

    # Service blank last
    service_sort = (1 if service else 2, service)

    return (
        category,
        service_sort,
        CONFIG.SEVERITY_ORDER.get(severity_label, 99),
        rec_id,
        account_id,
        resource
    )


def _write_excel_headers(ws, headers: List[str]) -> None:
    """Write headers to Excel worksheet."""
    header_fill = PatternFill(
        start_color=CONFIG.EXCEL_HEADER_COLOR,
        end_color=CONFIG.EXCEL_HEADER_COLOR,
        fill_type="solid"
    )
    header_font = Font(bold=True, color=CONFIG.EXCEL_HEADER_TEXT_COLOR)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill


def _write_recommendation_row(ws, row_num: int, rec: Dict) -> None:
    """Write a single recommendation row to Excel worksheet."""
    # Column mapping: Section, Policy, Severity, Account, Account Name,
    # Status, Resource, First Seen, Remediation, Docs, Tags
    ws.cell(row=row_num, column=1, value=rec.get('CATEGORY', ''))
    ws.cell(row=row_num, column=2, value=rec.get('TITLE', ''))

    # Severity as label
    severity = rec.get('SEVERITY', '')
    ws.cell(row=row_num, column=3, value=CONFIG.SEVERITY_MAP.get(severity, f'Unknown ({severity})'))

    ws.cell(row=row_num, column=4, value=rec.get('ACCOUNT_ID', ''))
    ws.cell(row=row_num, column=5, value=rec.get('ACCOUNT_ALIAS', ''))

    # Format status for readability
    status = rec.get('STATUS', '')
    ws.cell(row=row_num, column=6, value=_format_status(status))

    # Individual resource
    ws.cell(row=row_num, column=7, value=rec.get('RESOURCE', ''))

    # First Seen date
    ws.cell(row=row_num, column=8, value=rec.get('FIRST_SEEN', ''))

    # Remediation text
    ws.cell(row=row_num, column=9, value=rec.get('REMEDIATION', ''))

    # Docs: hyperlink to Fortinet docs with policy ID as link text
    rec_id = rec.get('REC_ID', '')
    if rec_id:
        rec_id_upper = rec_id.upper().replace('-', '_')
        url = CONFIG.POLICY_DOCS_URL.format(policy_id=rec_id_upper)
        cell = ws.cell(row=row_num, column=10)
        cell.value = rec_id
        cell.hyperlink = url
        cell.font = Font(color=CONFIG.EXCEL_LINK_COLOR, underline="single")
    else:
        ws.cell(row=row_num, column=10, value='')

    # Tags
    ws.cell(row=row_num, column=11, value=rec.get('TAGS', ''))


def _format_status(status: str) -> str:
    """Format status value for better readability."""
    if not status:
        return ''
    # Handle known statuses explicitly, fall back to PascalCase splitting
    status_map = {
        'noncompliant': 'Non Compliant',
        'compliant': 'Compliant',
        'couldnotassess': 'Could Not Assess',
        'suppressed': 'Suppressed',
    }
    mapped = status_map.get(status.lower())
    if mapped:
        return mapped
    # Insert space before capital letters for unknown PascalCase values
    return re.sub(r'(?<!^)(?=[A-Z])', ' ', status)


def _adjust_column_widths(ws, num_columns: int) -> None:
    """Auto-adjust column widths in worksheet."""
    for col_num in range(1, num_columns + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, CONFIG.EXCEL_MAX_COLUMN_WIDTH)
        ws.column_dimensions[column].width = adjusted_width


# ============================================================================
# Report Concatenation
# ============================================================================

def concatenate_reports(
    report_dir: str,
    output_format: OutputFormat,
    output_file: Optional[str],
    verbose: bool,
    include_compliant: bool = False,
    env: Dict[str, str] = None,
    cloud_type: CloudType = None,
    skip_tags: bool = False,
    report_name: str = ''
) -> str:
    """Concatenate individual report files into a single consolidated report."""
    Logger.verbose("Collecting recommendations from all accounts...", verbose)

    report_files = _find_report_files(report_dir)
    if not report_files:
        raise FileNotFoundError(f"No JSON report files found in: {report_dir}")

    Logger.verbose(f"Found {len(report_files)} report file(s)", verbose)

    # Extract metadata from first file
    metadata = _extract_metadata(report_files[0])
    Logger.verbose(f"Report: {metadata['title']}", verbose)
    Logger.verbose(f"Type: {metadata['type']}", verbose)
    Logger.verbose(f"Time: {metadata['time']}", verbose)

    # Collect all recommendations and calculate summary
    all_recommendations, all_accounts = _collect_recommendations(report_files, verbose)
    total_recommendations = len(all_recommendations)
    Logger.verbose(f"Total recommendations collected: {total_recommendations}", verbose)

    # Enrich recommendations with policy metadata (remediation text)
    if env and output_format != OutputFormat.JSON:
        Logger.info("Fetching policy metadata for remediation details...")
        all_policies = _fetch_all_policies(env, verbose)
        if all_policies:
            for rec in all_recommendations:
                rec_id = rec.get('REC_ID', '')
                if rec_id and not rec.get('REMEDIATION'):
                    policy = all_policies.get(rec_id, {})
                    rec['REMEDIATION'] = policy.get('remediation', '')

    # Fetch resource tags from inventory API
    tags_lookup = {}
    if env and cloud_type and not skip_tags and output_format != OutputFormat.JSON:
        urns = _collect_violation_urns(all_recommendations)
        if urns:
            Logger.info(f"Fetching tags for {len(urns)} unique resources from inventory...")
            tags_lookup = _fetch_inventory_tags(urns, cloud_type, env, verbose)
            Logger.info(f"Retrieved tags for {len(tags_lookup)} resources")
        else:
            Logger.verbose("No ARN resources found in violations, skipping tag fetch", verbose)

    # Enrich violations with first-seen dates from history
    if report_name and output_format != OutputFormat.JSON:
        _enrich_with_first_seen(all_recommendations, report_name, verbose)

    Logger.verbose("Calculating overall summary...", verbose)
    summary = _calculate_aggregate_summary(report_files)

    # Create consolidated data
    consolidated_data = {
        'reportTitle': metadata['title'],
        'reportType': metadata['type'],
        'reportTime': metadata['time'],
        'recommendations': all_recommendations,
        'summary': [summary]
    }

    # Determine output file (default to output directory)
    output_file = _determine_output_file(output_file, output_format, report_dir)
    _ensure_output_directory(output_file, verbose)

    Logger.verbose(f"Output file: {output_file}", verbose)

    # Generate output
    if output_format == OutputFormat.JSON:
        _write_json_output(consolidated_data, output_file, verbose)
    else:
        _write_excel_output(consolidated_data, output_file, verbose, include_compliant, tags_lookup)

    # Print summary
    _print_concatenation_summary(all_accounts, total_recommendations, output_file)

    return report_dir


def _find_report_files(report_dir: str) -> List[Path]:
    """Find all JSON report files in directory."""
    report_files = []
    for file in sorted(Path(report_dir).glob('*.json')):
        if file.name != 'all_accounts.json':  # Exclude previously generated files
            report_files.append(file)
    return report_files


def _extract_metadata(report_file: Path) -> Dict[str, str]:
    """Extract metadata from first report file."""
    with open(report_file, 'r') as f:
        data = json.load(f)
    return {
        'title': data.get('reportTitle', ''),
        'type': data.get('reportType', ''),
        'time': data.get('reportTime', '')
    }


def _collect_recommendations(report_files: List[Path], verbose: bool) -> Tuple[List[Dict], List[str]]:
    """Collect all recommendations from report files."""
    all_recommendations = []
    all_accounts = []
    
    for report_file in report_files:
        with open(report_file, 'r') as f:
            data = json.load(f)
        
        account_id = data.get('accountId', '')
        account_alias = data.get('accountAlias', '')
        all_accounts.append(account_id)
        
        Logger.verbose(f"Processing account: {account_id} ({account_alias})", verbose)
        
        recommendations = data.get('recommendations', [])
        all_recommendations.extend(recommendations)
    
    return all_recommendations, all_accounts


def _calculate_aggregate_summary(report_files: List[Path]) -> Dict:
    """Calculate aggregate summary statistics from all reports."""
    summary_fields = [
        'ASSESSED_RESOURCE_COUNT', 'NUM_COMPLIANT', 'NUM_NOT_COMPLIANT',
        'NUM_RECOMMENDATIONS', 'NUM_SEVERITY_1_NON_COMPLIANCE',
        'NUM_SEVERITY_2_NON_COMPLIANCE', 'NUM_SEVERITY_3_NON_COMPLIANCE',
        'NUM_SEVERITY_4_NON_COMPLIANCE', 'NUM_SEVERITY_5_NON_COMPLIANCE',
        'NUM_SUPPRESSED', 'SUPPRESSED_RESOURCE_COUNT', 'VIOLATED_RESOURCE_COUNT'
    ]
    
    totals = {field: 0 for field in summary_fields}
    
    for report_file in report_files:
        with open(report_file, 'r') as f:
            data = json.load(f)
        
        summary = data.get('summary', [{}])[0]
        for field in summary_fields:
            totals[field] += summary.get(field, 0)
    
    return totals


def _determine_output_file(output_file: Optional[str], output_format: OutputFormat, report_dir: str) -> str:
    """Determine output file path."""
    if output_file:
        return output_file
    
    # Default to output directory
    output_dir = CONFIG.OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)
    
    if output_format == OutputFormat.JSON:
        return os.path.join(output_dir, CONFIG.DEFAULT_JSON_OUTPUT)
    return os.path.join(output_dir, CONFIG.DEFAULT_EXCEL_OUTPUT)


def _ensure_output_directory(output_file: str, verbose: bool) -> None:
    """Ensure output directory exists."""
    output_dir = os.path.dirname(output_file)
    if output_dir and output_dir not in ['.', './']:
        os.makedirs(output_dir, exist_ok=True)
        Logger.verbose(f"Created output directory: {output_dir}", verbose)


def _write_json_output(data: Dict, output_file: str, verbose: bool) -> None:
    """Write consolidated data as JSON."""
    Logger.verbose("Creating single concatenated report file (JSON)...", verbose)
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2)
    Logger.info(f"Saved concatenated report to: {output_file}")


def _write_excel_output(data: Dict, output_file: str, verbose: bool, include_compliant: bool = False, tags_lookup: Dict = None) -> None:
    """Write consolidated data as Excel."""
    Logger.verbose("Creating Excel spreadsheet...", verbose)
    create_excel_from_report(data, output_file, include_compliant, tags_lookup)
    Logger.info(f"Saved Excel spreadsheet to: {output_file}")


def _print_concatenation_summary(accounts: List[str], recommendations: int, output_file: str) -> None:
    """Print concatenation summary."""
    print()
    Logger.info("Summary:")
    Logger.info(f"  Processed {len(accounts)} account(s)")
    Logger.info(f"  Total recommendations: {recommendations}")
    print()
    Logger.info(f"Output file: {output_file}")


def investigate_account(data: Dict, account_id: str) -> None:
    """Investigate 'Could Not Assess' issues for a specific account."""
    recommendations = data.get('recommendations', [])
    
    # Filter for the specific account
    account_recs = [rec for rec in recommendations if rec.get('ACCOUNT_ID', '') == account_id]
    
    if not account_recs:
        Logger.warning(f"No recommendations found for account: {account_id}")
        return
    
    # Filter for "Could Not Assess" entries
    could_not_assess = [rec for rec in account_recs if rec.get('STATUS', '').upper() == 'COULDNOTASSESS']
    
    print()
    Logger.info(f"Investigation Report for Account: {account_id}")
    Logger.info("=" * 70)
    print()
    
    Logger.info(f"Total recommendations for this account: {len(account_recs)}")
    Logger.info(f"'Could Not Assess' entries: {len(could_not_assess)}")
    print()
    
    if not could_not_assess:
        Logger.info("No 'Could Not Assess' issues found for this account.")
        return
    
    # Group by service and policy
    by_service = {}
    by_policy = {}
    
    for rec in could_not_assess:
        service = rec.get('SERVICE', 'N/A')
        policy_id = rec.get('REC_ID', 'N/A')
        policy_title = rec.get('TITLE', 'N/A')
        
        if service not in by_service:
            by_service[service] = []
        by_service[service].append(rec)
        
        if policy_id not in by_policy:
            by_policy[policy_id] = {
                'title': policy_title,
                'service': service,
                'count': 0
            }
        by_policy[policy_id]['count'] += 1
    
    # Show summary by service
    Logger.info("Breakdown by Service:")
    for service, recs in sorted(by_service.items()):
        Logger.info(f"  {service}: {len(recs)} policy(ies)")
    print()
    
    # Show detailed policy list
    Logger.info("Affected Policies:")
    for policy_id, info in sorted(by_policy.items(), key=lambda x: x[1]['count'], reverse=True):
        Logger.info(f"  [{info['count']} occurrence(s)] {policy_id}")
        Logger.info(f"    Service: {info['service']}")
        Logger.info(f"    Policy: {info['title']}")
        Logger.info(f"    Resource Count: {sum(r.get('RESOURCE_COUNT', 0) for r in could_not_assess if r.get('REC_ID') == policy_id)}")
        Logger.info(f"    Assessed Count: {sum(r.get('ASSESSED_RESOURCE_COUNT', 0) for r in could_not_assess if r.get('REC_ID') == policy_id)}")
        print()
    
    # Common causes analysis
    Logger.info("Possible Causes:")
    Logger.info("  1. Missing IAM Permissions: Lacework integration role may lack permissions")
    Logger.info("     to access certain AWS services or resources in this account.")
    Logger.info("  2. Service Not Enabled: Required AWS services may not be enabled")
    Logger.info("     (e.g., IAM Access Analyzer, Config, etc.)")
    Logger.info("  3. Resource Not Available: The resource type may not exist in this account.")
    Logger.info("  4. Integration Issues: Lacework may not be properly integrated with this account.")
    print()
    Logger.info("Recommendations:")
    Logger.info("  - Check Lacework integration status for this account")
    Logger.info("  - Verify IAM permissions for Lacework integration role")
    Logger.info("  - Ensure required AWS services are enabled")
    Logger.info("  - Review account configuration in Lacework console")
    print()


def _load_consolidated_data(report_output_dir: str, verbose: bool) -> Optional[Dict]:
    """Load consolidated data from individual report files for investigation."""
    report_files = []
    for file in sorted(Path(report_output_dir).glob('*.json')):
        if file.name != 'all_accounts.json':
            report_files.append(file)
    if not report_files:
        return None
    all_recommendations, _ = _collect_recommendations(report_files, verbose)
    metadata = _extract_metadata(report_files[0])
    summary = _calculate_aggregate_summary(report_files)
    return {
        'reportTitle': metadata['title'],
        'reportType': metadata['type'],
        'reportTime': metadata['time'],
        'recommendations': all_recommendations,
        'summary': [summary]
    }


# ============================================================================
# Main Function
# ============================================================================

def main() -> None:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Get compliance reports for all accounts in a Lacework instance',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s api-key/api-key.json "CIS Amazon Web Services Foundations Benchmark v1.4.0"
  %(prog)s api-key/api-key.json "My Custom AWS Framework" --cloud-type aws
  %(prog)s api-key/api-key.json "CIS Microsoft Azure Foundations Benchmark v1.5.0" -v
  %(prog)s api-key/api-key.json "CIS Amazon Web Services Foundations Benchmark v1.4.0" -f json
  %(prog)s api-key/api-key.json "CIS Amazon Web Services Foundations Benchmark v1.4.0" -o my-report.xlsx
  %(prog)s api-key/api-key.json "CIS Amazon Web Services Foundations Benchmark v1.4.0" --no-concatenate

To find available compliance reports:
  lacework report-definitions list                    (may not show custom frameworks)

For custom frameworks, specify --cloud-type explicitly:
  %(prog)s api-key/api-key.json "My Custom Report" --cloud-type aws
        """
    )
    
    parser.add_argument('api_key_path', help='Path to the Lacework API key JSON file')
    parser.add_argument('report_name', help='Name of the compliance report to fetch')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output')
    parser.add_argument('--use-cache', action='store_true', help='Use cached account list data')
    parser.add_argument('--no-concatenate', action='store_true', help='Skip concatenation (only save individual JSON files)')
    parser.add_argument('--keep-intermediate', action='store_true', help='Keep intermediate JSON files in output/ directory after concatenation')
    parser.add_argument('-f', '--format', choices=['json', 'excel'], default='excel', help='Output format for concatenation (default: excel)')
    parser.add_argument('-o', '--output', help='Output file path for concatenated report')
    parser.add_argument('--cloud-type', choices=['aws', 'azure', 'gcp'],
                        help='Cloud type override. Required for custom frameworks not listed in report-definitions.')
    parser.add_argument('--include-compliant', action='store_true',
                        help='Include compliant policies (no violations) in the Excel output')
    parser.add_argument('--test', action='store_true',
                        help='Test mode: limit to first 3 accounts')
    parser.add_argument('--skip-tags', action='store_true',
                        help='Skip fetching resource tags from inventory API')
    parser.add_argument('--investigate-account', help='Investigate "Could Not Assess" issues for a specific account ID')
    parser.add_argument('--use-reports-api', action='store_true',
                        help='Use the /api/v2/Reports endpoint instead of LQL queries (legacy; Reports API may return 500)')
    
    args = parser.parse_args()
    
    # Convert format string to enum
    output_format = OutputFormat.JSON if args.format == 'json' else OutputFormat.EXCEL
    check_required_tools(output_format)
    
    Logger.info("Starting compliance report collection...")
    Logger.info(f"Report: {args.report_name}")
    Logger.info(f"API Key: {args.api_key_path}")
    
    # Load and configure API key
    api_key = load_api_key(args.api_key_path)
    env = configure_lacework(api_key, args.verbose)
    
    # Get report type (cloud type)
    if args.cloud_type:
        cloud_type = CloudType(args.cloud_type)
        Logger.info(f"Using specified cloud type: {cloud_type.value}")
    else:
        Logger.info("Validating report and determining cloud type...")
        cloud_type = get_report_info(args.report_name, env, args.verbose)
        if not cloud_type:
            Logger.error(
                f"Failed to determine cloud type for report '{args.report_name}'. "
                "This may be a custom framework not listed in report-definitions. "
                "Try specifying --cloud-type aws|azure|gcp explicitly."
            )
            sys.exit(1)
        Logger.info(f"Report type: {cloud_type.value}")
    
    # Get accounts list
    Logger.info(f"Fetching list of {cloud_type.value} accounts...")
    accounts = get_accounts(cloud_type, env, args.verbose, args.use_cache, CONFIG.CACHE_DIR)
    
    if not accounts:
        Logger.warning(f"No accounts found for cloud type: {cloud_type.value}")
        sys.exit(0)
    
    if args.test:
        accounts = accounts[:3]
        Logger.info(f"Test mode: limiting to first {len(accounts)} account(s)")

    account_count = len(accounts)
    Logger.info(f"Found {account_count} account(s)")
    
    # Create output directory (sanitize report name to prevent path traversal)
    safe_report_name = args.report_name.replace('/', '_').replace('\\', '_').replace('..', '_')
    report_output_dir = os.path.join(CONFIG.OUTPUT_DIR, f"{safe_report_name}_{cloud_type.value}")
    os.makedirs(report_output_dir, exist_ok=True)
    Logger.info(f"Output directory: {report_output_dir}")
    
    # Fetch reports
    use_reports_api = args.use_reports_api

    if not use_reports_api:
        # Default: try LQL policy queries first
        result = _fetch_reports_via_lql(
            args.report_name, cloud_type, accounts,
            report_output_dir, env, args.verbose, args.test
        )
        if result is None:
            # Report definition not found — silently fall back to Reports API
            use_reports_api = True
        else:
            success_count, failure_count = result

    if use_reports_api:
        # Fetch report per account via /api/v2/Reports
        Logger.verbose("Using Reports API", args.verbose)
        success_count = 0
        failure_count = 0

        for account_num, account in enumerate(accounts, 1):
            Logger.info(f"[{account_num}/{account_count}] Processing account: {account}")

            safe_account_name = account.replace('/', '_').replace(':', '_')
            output_file = os.path.join(report_output_dir, f"{safe_account_name}.json")

            if get_report_for_account(cloud_type, args.report_name, account, output_file, env, args.verbose):
                success_count += 1
            else:
                failure_count += 1
                Logger.warning(f"Failed to get report for account: {account}")
                if account_num == 1:
                    Logger.error(
                        f"First account returned no data. Report name '{args.report_name}' "
                        "may not exist."
                    )
                    sys.exit(1)

            time.sleep(CONFIG.REQUEST_DELAY)

    # Summary
    print()
    Logger.info("Summary:")
    Logger.info(f"  Total accounts: {len(accounts) if use_reports_api else success_count}")
    Logger.info(f"  Successful: {success_count}")
    Logger.info(f"  Failed: {failure_count}")
    if args.no_concatenate:
        Logger.info(f"  Individual reports saved to: {report_output_dir}")
    
    if failure_count > 0:
        sys.exit(1)
    
    # Concatenate reports if requested (default is True)
    consolidated_data = None
    if not args.no_concatenate:
        print()
        Logger.info("Concatenating reports...")
        try:
            concatenate_reports(report_output_dir, output_format, args.output, args.verbose,
                                args.include_compliant, env=env, cloud_type=cloud_type,
                                skip_tags=args.skip_tags, report_name=args.report_name)
            Logger.verbose("Successfully created concatenated report", args.verbose)
            
            # Load consolidated data for investigation if needed (before cleanup)
            if args.investigate_account:
                consolidated_data = _load_consolidated_data(report_output_dir, args.verbose)
            
            # Clean up intermediate files unless --keep-intermediate is specified
            if not args.keep_intermediate:
                Logger.verbose("Cleaning up intermediate files...", args.verbose)
                try:
                    shutil.rmtree(report_output_dir)
                    Logger.verbose(f"Removed directory: {report_output_dir}", args.verbose)
                    Logger.verbose("Cleaned up intermediate JSON files", args.verbose)
                except Exception as e:
                    Logger.warning(f"Failed to clean up intermediate files: {e}")
        except Exception as e:
            Logger.error(f"Failed to concatenate reports: {e}")
            Logger.warning("Intermediate files preserved for debugging")
            sys.exit(1)
    elif args.investigate_account:
        consolidated_data = _load_consolidated_data(report_output_dir, args.verbose)
    
    # Run investigation if requested
    if args.investigate_account:
        if not consolidated_data:
            Logger.error("Cannot investigate account: No consolidated data available. Run without --no-concatenate or ensure intermediate files exist.")
            sys.exit(1)
        investigate_account(consolidated_data, args.investigate_account)


if __name__ == '__main__':
    main()
